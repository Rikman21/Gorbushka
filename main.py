import asyncio
import logging
import json
import urllib.parse
import time
import os
from aiohttp import web
from openpyxl import Workbook, load_workbook # <--- Добавили load_workbook
from io import BytesIO

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import WebAppInfo, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile

import database 

TOKEN = "8516086910:AAFugoM9-OjnOOJFT3flpcyUOhh4P9alxSY" # <--- НЕ ЗАБУДЬТЕ ВЕРНУТЬ ТОКЕН
WEB_APP_URL = "https://rikman21.github.io/Gorbushka/" 

logging.basicConfig(level=logging.INFO)
bot = Bot(token=TOKEN)
dp = Dispatcher()

# --- ФЕЙКОВЫЙ СЕРВЕР ---
async def health_check(request): return web.Response(text="Alive")
async def start_dummy_server():
    port = int(os.environ.get("PORT", 8080))
    app = web.Application()
    app.router.add_get('/', health_check)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', port)
    await site.start()

# --- ГЕНЕРАЦИЯ ШАБЛОНА ---
def generate_excel_template():
    products = database.get_catalog_for_excel()
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"
    # SKU в 1 колонке, Цена в 6 колонке (индексы A и F)
    headers = ["SKU (Не менять!)", "Модель", "Память", "Цвет", "Сим", "ВАША ЦЕНА (Рубли)"]
    ws.append(headers)
    for p in products:
        row = list(p) + [""] 
        ws.append(row)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()

# --- ЛОГИКА ЗАГРУЗКИ ЦЕН ---
@dp.message(F.document)
async def handle_document(message: types.Message):
    # Проверка: это Excel?
    if not message.document.file_name.endswith('.xlsx'):
        return await message.answer("❌ Это не Excel. Пришлите файл .xlsx")

    user_id = message.from_user.id
    username = message.from_user.username or "Продавец"
    
    wait_msg = await message.answer("⏳ Скачиваю и обрабатываю прайс...")

    try:
        # 1. Скачиваем файл в память
        bot_file = await bot.get_file(message.document.file_id)
        file_data = await bot.download_file(bot_file.file_path)
        
        # 2. Открываем через openpyxl
        wb = load_workbook(file_data)
        ws = wb.active
        
        # 3. Читаем строки (пропускаем заголовок)
        prices_to_update = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row[0] = SKU (Колонка A)
            # row[5] = ЦЕНА (Колонка F)
            sku = row[0]
            price_raw = row[5]
            
            # Чистим цену (если там пробелы или текст)
            price = None
            if price_raw:
                try:
                    price = int(str(price_raw).replace(" ", "").replace("₽", ""))
                except:
                    price = None # Если написали бред, считаем что цены нет
            
            if sku: # Если SKU есть, добавляем в список на обработку
                prices_to_update.append((sku, price))
        
        # 4. Пишем в базу
        updated_count = database.update_prices_from_excel(user_id, username, prices_to_update)
        
        await wait_msg.edit_text(f"✅ **Прайс обновлен!**\n\nТоваров в продаже: {updated_count}\n\nТеперь они видны в поиске.")
        
    except Exception as e:
        logging.error(e)
        await wait_msg.edit_text("❌ Ошибка при чтении файла. Убедитесь, что вы не меняли колонку SKU.")

# --- СТАРТ И WEBAPP ---
@dp.message(Command("start"))
async def start(message: types.Message):
    user_id = message.from_user.id
    offers_list = database.get_all_offers_for_web()
    offers_json = json.dumps(offers_list)
    offers_encoded = urllib.parse.quote(offers_json)
    timestamp = int(time.time())
    full_url = f"{WEB_APP_URL}?data={offers_encoded}&ver={timestamp}&uid={user_id}"

    kb = [[KeyboardButton(text="📱 ОТКРЫТЬ МАРКЕТ", web_app=WebAppInfo(url=full_url))]]
    await message.answer("👋 Горбушка Онлайн\n\nПродавцы: загружайте прайс Excel прямо сюда файлом.", reply_markup=ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True))

@dp.message(F.web_app_data)
async def handle_webapp(message: types.Message):
    data = message.web_app_data.data
    user_id = message.chat.id
    username = message.from_user.username or "Клиент"

    if data == "REQ_TEMPLATE":
        file_bytes = generate_excel_template()
        document = BufferedInputFile(file_bytes, filename="Gorbushka_Price_Template.xlsx")
        await message.answer_document(document, caption="📉 **Ваш шаблон**\n1. Заполните цены.\n2. Пришлите этот файл мне в чат.")
        return

    if data.startswith("REQ_BUY"):
        parts = data.split("|")
        seller_id = int(parts[1])
        product_name = parts[3]
        price = parts[4]
        
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ В наличии", callback_data=f"confirm_{user_id}")],
            [InlineKeyboardButton(text="❌ Нет", callback_data=f"reject_{seller_id}")]
        ])
        try:
            await bot.send_message(seller_id, f"🔔 <b>ЗАКАЗ!</b>\n\n📦 {product_name}\n💰 {price}р\n👤 @{username}", reply_markup=kb, parse_mode="HTML")
            await message.answer("⏳ Запрос отправлен продавцу...")
        except:
            await message.answer("Продавец не найден.")

    # Старый метод (поштучно) можно оставить или убрать, он не мешает
    elif data.startswith("NEW_PRICE"):
        await message.answer("⚠️ Лучше используйте загрузку через Excel для надежности.")

@dp.callback_query(F.data.startswith("confirm_"))
async def confirm_order(callback: types.CallbackQuery):
    buyer_id = int(callback.data.split("_")[1])
    await callback.message.edit_text(f"✅ Подтверждено!", reply_markup=None)
    await bot.send_message(buyer_id, f"🎉 Продавец подтвердил!\nКонтакт: @{callback.from_user.username}")

@dp.callback_query(F.data.startswith("reject_"))
async def reject_order(callback: types.CallbackQuery):
    await callback.message.edit_text(f"🚫 Отказ.", reply_markup=None)

async def main():
    database.init_db()
    await start_dummy_server()
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
